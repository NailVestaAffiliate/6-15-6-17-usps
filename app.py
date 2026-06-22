"""
NailVesta 達人補發追蹤 — Streamlit 改良版
執行：streamlit run app.py
存檔：有設定 Google Sheets 就寫雲端、多人共用；沒設定就寫本機 CSV。
"""
import io
import os
import pandas as pd
import streamlit as st

st.set_page_config(page_title="達人補發追蹤", page_icon="💅", layout="wide")

SAVE_FILE = "達人補發追蹤_state.csv"

# 狀態流程（emoji 當顏色用，一眼看出卡在哪一步）
STATUS = ["⚪ 待達人回報", "🟡 達人已通知", "🔵 已補發", "🟢 已送達結案"]
DEFAULT_STATUS = "🟡 達人已通知"   # 此清單都是已回報未收到，預設已通知

# USPS 官方追蹤網址
USPS_BASE = "https://tools.usps.com/go/TrackConfirmAction?tLabels="
def track_url(num):
    n = str(num).replace(" ", "").strip()
    return USPS_BASE + n if n and n.lower() != "nan" else ""

# 達人, 日期, 姓名, Email, 電話, 地址, 款式, 尺寸, 原單號(LV/HB), 原 USPS Tracking
RAW = [
 ("xiomaradenise_","2026/06/15","Xiomara DiGiovanna","xiomaradigiovanna@gmail.com","8654123598","1748 Quail Ridge Loop, Kissimmee, FL 34744","Pastel Coast, Sunlit Petals, Aloha Bloom","M","LV068190166US","9300110993319120459047"),
 ("uhhhnegin-4","2026/06/15","Negin Mandavi","neginmandavi@gmail.com","8912354789","15950 Chase Hill Blvd, San Antonio, TX 78256","Berry Bowtie, Mochi Blossom","S","LV068190197US","9300110993319120459061"),
 ("anaiddm","2026/06/15","Diana Medina","Anaiddm13@gmail.com","8654932147","10223 San Luis Ave, South Gate, CA 90280","Fairy Nectar, Pastel Bloom, Coraline Glow","S","LV068190152US","9300110993319120459023"),
 ("akroulette","2026/06/15","Jamal Wint","Akroulettepromotions@gmail.com","5214896325","745 Scenic Hwy Apt 5208, Lawrenceville, GA 30045","Aqua Blush, Golden Hibiscus, Petal Throne","M","LV068190206US","9300110993319120459078"),
 ("life_witheb","2026/06/15","Eboni Logan","lifewitheb01@gmail.com","1547896325","3123 Calvary Drive Apt B3, Raleigh, NC 27604","Rosé Petal, Seaside Sundae, Tidal Flower","M","LV068190183US","9300110993319120459054"),
 ("keanna.edit","2026/06/15","Keanna Barocas","keanna.ugc@gmail.com","4589312475","14824 Faversham Cir, Orlando, FL 32826","Rouge Letter, Teal Blossom, Aloha Bloom","L","LV068190170US","9300110993319120459030"),
 ("avibeinpisces","2026/06/15","Salina Spain","salina.spain@outlook.com","8654125983","14624 Fieldcrest Lane, Burnsville, MN 55306","Teal Blossom, Jade Garden, Petal Throne","M","LV068190245US","9300110993319120459115"),
 ("glambeatzbyzee","2026/06/15","Zatese Brown","makeupbyzatese@gmail.com","9563214789","5001 Cypress Creek Ave E Apt 1010, Tuscaloosa, AL 35405","Jade Garden, Mochi Blossom","M","LV068190210US","9300110993319120459085"),
 ("saydizzle-2","2026/06/15","Sada Policar","Hello@sadaarika.com","4589632147","2020 Barranca St Unit 410, Los Angeles, CA 90031","Fairy Nectar, Champagne Blossom, Jade Garden","S","LV068190237US","9300110993319120459108"),
 ("cyaira.a","2026/06/15","Cyaira Angeliq","cyairaangeliq@gmail.com","2679920486","650 Kerper St, Philadelphia, PA 19111","Aqua Blush, Cowgirl Charm, Pearl Tide","S","LV068190223US","9300110993319120459092"),
 ("findsbykayla","2026/06/16","Kayla Ring","kaylanicolering@gmail.com","612-978-7166","17619 Grant Street NW, Elk River, MN 55330","Island Paradise, Citrus Veil, Tidal Flower","S","LV068316224US","9300110993319120584749"),
 ("shelives.on.whimsylane","2026/06/16","Mia Walker","knottycroture@gmail.com","5488641235","8958 River Island Drive Apt 203, Savage, MD 20763","Apricot Cream, Rosé Angel, Marine Glow","M","LV068316255US","9300110993319120584763"),
 ("samantharose.__","2026/06/16","Samantha Rosario","Ssamantharosee6@gmail.com","6587423915","211 Meade St, Perth Amboy, NJ 08861","Acai Bloom, Seaside Sundae, Sunflower Safari","M","LV068316215US","9300110993319120584725"),
 ("clarityhome","2026/06/16","Clarity Home","clarityhomeco@gmail.com","8456912354","8035 100th Ct, Vero Beach, FL 32967","Mint Petal, Silk Blossom","S","LV068316241US","9300110993319120584732"),
 ("skyepalafoxx","2026/06/16","Skyleynn Palafox","skyepalafox@gmail.com","1254789632","10809 Colony Wood Pl, Spring, TX 77380","Coraline Glow, Mochi Blossom, Petal Throne","M","LV068316238US","9300110993319120584756"),
 ("tutugoldd_","2026/06/16","Jade Rivera","Tutugoldd@gmail.com","4562157893","1892 English Court, Virginia Beach, VA 23454","Peach Ember, Tidal Flower, Sunflower Safari","L","LV068316269US","9300110993319120584770"),
 ("lindseyingleee","2026/06/17","Lindsey Hedrick","lindseyingleeee@gmail.com","4589623157","1301 S Park St Trlr C-39, Sapulpa, OK 74066","Royal Treasure, Jade Blossom, Papaya Bloom","L","HB089165365US","9300110663319120128877"),
 ("grwmm.gabi-5","2026/06/17","Esperanza Reyes","grwmm.gabi@gmail.com","912-331-9345","33 Villa St Apt 5, Douglas, GA 31533","Royal Treasure, Seaside Sundae, Tropical Breeze","M","HB089165405US","9300110663319120128914"),
 ("lovebeaches1","2026/06/17","Chelsea Chimere Brown","Chichibad.nco@gmail.com","5412863459","1386 Oakdale Ave Apt G, El Cajon, CA 92021","Citrus Daisy, Petal Throne, Pearl Tide","L","HB089165374US","9300110663319120128884"),
 ("tyetaughtuwell_","2026/06/17","Tyesha Draughn","tyetaughtuwell@gmail.com","5846231457","320 Eliza Way, Winterville, NC 28590","Sunset Treasure, Ocean Picnic, Honey Petal","M","HB089165388US","9300110663319120128891"),
 ("misamoresx3","2026/06/17","Melissa Penaloza","mellypcollabz@gmail.com","4521896325","3628 Cardamon Dr, Bakersfield, CA 93309","Fairy Garden, Island Paradise, Opal Dynasty","S","HB089165391US","9300110663319120128907"),
 ("prestonandtina7","2026/06/17","Tina Vielot","Prestonandtina7@gmail.com","4521369874","14261 Heritage Landing Blvd Apt 1627, Punta Gorda, FL 33955","Tropical Breeze, Pastel Coast, Golden Hibiscus","M","HB089165414US","9300110663319120128921"),
 ("allthingssdee","2026/06/17","Diamond Gray","diamondhen25@gmail.com","16562421138","809 Alpine Ct, Kissimmee, FL 34758","Rosé Petal, Fairy Garden","M","HB089165431US","9300110663319120128945"),
 ("dannastewart","2026/06/17","D'anna Stewart","hello.dannastewart@gmail.com","8956471235","1226 Tolkien Rd, Riverside, CA 92506","Ruby Bloom, Fairy Garden, Jade Garden","L","HB089165428US","9300110663319120128938"),
 ("candylites","2026/06/17","Elba Aguirre","elbadarling@gmail.com","4856789231","1557 Flamingo St, Beaumont, CA 92223","Royal Treasure, Petal Throne","M","HB089165459US","9300110663319120128969"),
 ("fggvron","2026/06/17","Veronica Ford","","12569321439","100 Edison Ave NW Apt 6302, Madison, AL 35757","Pastel Coast, Golden Hibiscus, Silk Blossom","M","HB089165445US","9300110663319120128952"),
 ("lexbillionzz-76","2026/06/17","Lex Billionz","Lexb.business@gmail.com","4589612354","3580 E Commerce Way, Sacramento, CA 95834","Prism Aura, Citrus Veil, Cowgirl Charm, Sunflower Safari","M","HB089165462US","9300110663319120128976"),
]

DATA_COLS = ["達人 Handle","日期","姓名","Email","電話","地址","款式","尺寸","原單號(LV/HB)","原 USPS Tracking"]
TRACK_COLS = ["狀態","補發日期","補發新單號","備註"]
COLS = DATA_COLS + TRACK_COLS


def base_df():
    df = pd.DataFrame(RAW, columns=DATA_COLS)
    df["狀態"] = DEFAULT_STATUS
    df["補發日期"] = ""
    df["補發新單號"] = ""
    df["備註"] = ""
    return df[COLS]


# ---------- 存檔：優先 Google Sheets，否則本機 CSV ----------
def _gsheet():
    """回傳 worksheet 物件；未設定或套件缺失則回傳 None。"""
    try:
        if "gcp_service_account" not in st.secrets:
            return None
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(st.secrets["gsheet_key"])
        return sh.sheet1
    except Exception:
        return None


def load_df():
    ws = _gsheet()
    if ws is not None:
        try:
            recs = ws.get_all_records()
            if recs:
                df = pd.DataFrame(recs).astype(str).fillna("")
                return _migrate(df)
        except Exception:
            pass
    if os.path.exists(SAVE_FILE):
        try:
            df = pd.read_csv(SAVE_FILE, dtype=str).fillna("")
            return _migrate(df)
        except Exception:
            pass
    return base_df()


def _migrate(saved):
    """固定資料(單號/地址/款式)一律取自內建 RAW；存檔只覆蓋進度欄，用 Handle 對應。"""
    base = base_df().set_index("達人 Handle")
    saved = saved.copy()

    # 舊版欄名 → 新版欄名
    rename = {"物流單號(LV/HB)": "原單號(LV/HB)", "物流單號 (LV/HB)": "原單號(LV/HB)",
              "USPS Tracking": "原 USPS Tracking"}
    saved = saved.rename(columns={k: v for k, v in rename.items() if k in saved.columns})

    # 舊版雙勾選欄 → 新狀態欄
    if "狀態" not in saved.columns and "我們是否補發" in saved.columns:
        def to_status(r):
            yes = lambda v: str(v).lower() in ("true", "1", "✓", "yes")
            if yes(r.get("我們是否補發", "")):
                return "🔵 已補發"
            if yes(r.get("達人是否通知", "")):
                return "🟡 達人已通知"
            return "⚪ 待達人回報"
        saved["狀態"] = saved.apply(to_status, axis=1)

    # 只把進度欄依 Handle 合併回 base
    if "達人 Handle" in saved.columns:
        saved = saved.set_index("達人 Handle")
        idx = base.index.intersection(saved.index)
        for c in ["狀態", "補發日期", "補發新單號", "備註"]:
            if c in saved.columns:
                base.loc[idx, c] = saved.loc[idx, c].fillna("")
    return base.reset_index()[COLS]


def save_df(df):
    df = df.astype(str)
    ws = _gsheet()
    if ws is not None:
        try:
            ws.clear()
            ws.update([df.columns.tolist()] + df.values.tolist())
            return "Google Sheets（雲端共用）"
        except Exception:
            pass
    df.to_csv(SAVE_FILE, index=False)
    return "本機 CSV"


def to_excel(df):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    color = {"⚪ 待達人回報": "F2F2F2", "🟡 達人已通知": "FFF2CC",
             "🔵 已補發": "DDEBF7", "🟢 已送達結案": "E2EFDA"}
    wb = Workbook(); ws = wb.active; ws.title = "達人補發追蹤"
    navy = PatternFill("solid", fgColor="1F3864")
    white = Font(name="Arial", bold=True, color="FFFFFF", size=10); base = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)
    ctr = Alignment("center", "center", wrap_text=True); lft = Alignment("left", "center", wrap_text=True)
    out_cols = ["序號"] + COLS
    ws.append(out_cols)
    for c in range(1, len(out_cols) + 1):
        cell = ws.cell(1, c); cell.fill = navy; cell.font = white; cell.alignment = ctr; cell.border = border
    for i, (_, r) in enumerate(df.iterrows(), start=1):
        ws.append([i] + [r[c] for c in COLS]); rn = i + 1
        fill = PatternFill("solid", fgColor=color.get(r["狀態"], "FFFFFF"))
        for c in range(1, len(out_cols) + 1):
            cell = ws.cell(rn, c); cell.font = base; cell.border = border; cell.fill = fill
            cell.alignment = ctr if out_cols[c-1] in ("序號","達人 Handle","尺寸","狀態") else lft
        for cn in ("電話","原單號(LV/HB)","原 USPS Tracking","補發新單號"):
            ws.cell(rn, out_cols.index(cn)+1).number_format = "@"
        # 單號加 USPS 超連結
        link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
        for cn in ("原單號(LV/HB)","原 USPS Tracking","補發新單號"):
            cell = ws.cell(rn, out_cols.index(cn)+1)
            url = track_url(cell.value)
            if url:
                cell.hyperlink = url; cell.font = link_font
    widths = [6,20,12,18,26,15,34,24,7,16,22,16,14,22,22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "B2"; ws.row_dimensions[1].height = 28
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ---------- UI ----------
st.title("💅 NailVesta 達人補發追蹤")
st.caption("未收到包裹的達人清單 · 2026/06/15–06/17 · 共 27 位")

if "df" not in st.session_state:
    st.session_state.df = load_df()
df = st.session_state.df

cnt = df["狀態"].value_counts()
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("達人總數", len(df))
c2.metric("⚪ 待回報", int(cnt.get("⚪ 待達人回報", 0)))
c3.metric("🟡 待補發", int(cnt.get("🟡 達人已通知", 0)))
c4.metric("🔵 已補發在途", int(cnt.get("🔵 已補發", 0)))
c5.metric("🟢 已結案", int(cnt.get("🟢 已送達結案", 0)))

f1, f2 = st.columns([3, 1])
kw = f1.text_input("🔍 搜尋（達人 / 姓名 / 款式 / 單號）", "")
flt = f2.selectbox("篩選狀態", ["全部"] + STATUS)

view = df.copy()
if kw:
    m = pd.Series(False, index=view.index)
    for col in ["達人 Handle", "姓名", "款式", "原單號(LV/HB)", "原 USPS Tracking", "補發新單號"]:
        m |= view[col].astype(str).str.contains(kw, case=False, na=False)
    view = view[m]
if flt != "全部":
    view = view[view["狀態"] == flt]

# 加入可點擊的 USPS 查詢連結欄（點單號直接開新分頁查件）
disp = view.copy()
disp["🔗 LV/HB 查詢"] = disp["原單號(LV/HB)"].map(track_url)
disp["🔗 USPS 查詢"] = disp["原 USPS Tracking"].map(track_url)
disp["🔗 補發查詢"] = disp["補發新單號"].map(track_url)

LINK_COLS = ["🔗 LV/HB 查詢", "🔗 USPS 查詢", "🔗 補發查詢"]
EDITABLE = ["狀態", "補發日期", "補發新單號", "備註"]

edited = st.data_editor(
    disp, use_container_width=True, hide_index=True, num_rows="fixed", key="editor",
    column_order=["狀態","達人 Handle","姓名","款式","尺寸",
                  "🔗 USPS 查詢","🔗 LV/HB 查詢","補發日期","補發新單號","🔗 補發查詢",
                  "備註","電話","Email","地址","日期","原單號(LV/HB)","原 USPS Tracking"],
    column_config={
        "狀態": st.column_config.SelectboxColumn("狀態", options=STATUS, width="medium", required=True),
        "補發日期": st.column_config.TextColumn("補發日期", help="YYYY/MM/DD", width="small"),
        "補發新單號": st.column_config.TextColumn("補發新單號", help="補發後新的 USPS 追蹤碼，填了下一欄自動產生查詢連結", width="medium"),
        "備註": st.column_config.TextColumn("備註", width="medium"),
        "款式": st.column_config.TextColumn("款式", width="large"),
        "地址": st.column_config.TextColumn("地址", width="large"),
        "🔗 USPS 查詢": st.column_config.LinkColumn("🔗 USPS 查詢", display_text=r"tLabels=(.+)", help="點擊查 USPS 物流", width="medium"),
        "🔗 LV/HB 查詢": st.column_config.LinkColumn("🔗 LV/HB 查詢", display_text=r"tLabels=(.+)", help="點擊查 LV/HB 單號", width="medium"),
        "🔗 補發查詢": st.column_config.LinkColumn("🔗 補發查詢", display_text=r"tLabels=(.+)", help="點擊查補發單號", width="medium"),
    },
    disabled=DATA_COLS + LINK_COLS,
)
# 只把可編輯欄位寫回主表（連結欄是即時計算，不回寫）
for c in EDITABLE:
    st.session_state.df.loc[edited.index, c] = edited[c]

b1, b2, b3 = st.columns(3)
if b1.button("💾 儲存進度", use_container_width=True, type="primary"):
    where = save_df(st.session_state.df)
    st.success(f"已儲存到：{where}")

xlsx = to_excel(st.session_state.df)
if xlsx is not None:
    b2.download_button("📥 匯出 Excel", xlsx, "達人補發追蹤.xlsx", use_container_width=True)
else:
    b2.download_button("📥 匯出 CSV",
                       st.session_state.df.to_csv(index=False).encode("utf-8-sig"),
                       "達人補發追蹤.csv", use_container_width=True)

if b3.button("🔄 重置", use_container_width=True):
    st.session_state.df = base_df()
    if os.path.exists(SAVE_FILE):
        os.remove(SAVE_FILE)
    st.rerun()

st.caption("狀態流程：⚪ 待達人回報 → 🟡 達人已通知（待補發）→ 🔵 已補發 → 🟢 已送達結案")
